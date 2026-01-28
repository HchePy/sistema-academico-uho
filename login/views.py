from django.shortcuts import render, redirect, get_object_or_404
from django.db import models

from .models import User, Estudiante, Profesor, Carrera, Materia, Nota, Horario, Noticia
from django.contrib.auth.decorators import login_required
from django.db.models import Avg, Q

import json
import os
from django.http import JsonResponse
from django.contrib.auth import get_user_model, authenticate, login, logout
from django.contrib import messages
from .forms import LoginForm, RegistroUsuarioForm, NoticiaForm, SubirNotasForm, SubirHorariosForm

import openpyxl

def logout_view(request):
    logout(request)
    return redirect('login')

def login_view(request):
    if request.user.is_authenticated:
        return redirect('base')
        
    form = LoginForm(request, data=request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            
            # Recordar Sesión
            if form.cleaned_data.get('remember'):
                request.session.set_expiry(1209600) # 2 semanas
            else:
                request.session.set_expiry(0) # Cerrar al cerrar navegador
            
            return redirect('base')
        else:
            messages.error(request, "Usuario o contraseña incorrectos")

    carreras = Carrera.objects.all()
    # Noticias públicas para el carrusel (visible para todos)
    noticias_globales = Noticia.objects.filter(visible_para='todos').order_by('-fecha_publicacion')[:3]
    
    # Formulario de Registro para el modal (si falló el registro, vendrá en context si usáramos la misma view,
    # pero como usamos una vista separada que redirige aquí en error, lo inicializamos vacío o con errores si redirigimos manualmente)
    # Por simplicidad, si hay errores en registro, registro_view renderizará login.html directamente.
    registro_form = RegistroUsuarioForm()
    
    return render(request, 'login.html', {
        'form': form,
        'registro_form': registro_form,
        'carreras': carreras,
        'noticias_globales': noticias_globales
    })

def registro_view(request):
    if request.method == 'POST':
        form = RegistroUsuarioForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, f"Usuario creado exitosamente. Rol: {'Profesor' if user.is_profesor else 'Estudiante'}")
            return redirect('login')
        else:
             # Si hay errores, renderizamos el login con el modal abierto y los errores
             messages.error(request, "Error en el registro. Verifique los campos.")
             login_form = LoginForm()
             carreras = Carrera.objects.all()
             noticias_globales = Noticia.objects.filter(visible_para='todos').order_by('-fecha_publicacion')[:3]
             return render(request, 'login.html', {
                 'form': login_form,
                 'registro_form': form,
                 'carreras': carreras,
                 'noticias_globales': noticias_globales,
                 'show_register_modal': True
             })
    return redirect('login')

@login_required
def base_view(request):
    if request.user.is_profesor:
        try:
            profesor = request.user.profesor
            materias = profesor.materias.all()
            
            # Metricas para el profesor
            total_alumnos = Estudiante.objects.filter(carrera=profesor.carrera).count()
            # 1. Noticias que ve:
            noticias_feed = Noticia.objects.filter(
                Q(visible_para='todos') |
                (Q(visible_para='profesores') & (Q(carrera=profesor.carrera) | Q(carrera__isnull=True)))
            ).order_by('-fecha_publicacion')[:5]
            
            # 2. Mis noticias
            mis_noticias = Noticia.objects.filter(autor=profesor).order_by('-fecha_publicacion')
            
            noticia_form = NoticiaForm()
            subir_notas_form = SubirNotasForm()
            subir_horarios_form = SubirHorariosForm()

            context = {
                'materias': materias,
                'carreras': Carrera.objects.all(),
                'noticias': noticias_feed,
                'mis_noticias': mis_noticias,
                'noticia_form': noticia_form,
                'subir_notas_form': subir_notas_form,
                'subir_horarios_form': subir_horarios_form,
                'metricas': {
                    'total_alumnos': total_alumnos,
                    'promedio_grupal': 7.5, 
                    'tareas_pendientes': 12
                }
            }

            # Logica extra para Jefe de Carrera (Carga de datos directa)
            if profesor.nivel == 'jefe de carrera':
                todos_profesores = Profesor.objects.all().select_related('user', 'carrera').prefetch_related('materias')
                lista_profesores = []
                for p in todos_profesores:
                    materias_ids = [m.id for m in p.materias.all()]
                    lista_profesores.append({
                        'obj': p, # Objeto completo por si acaso
                        'id': p.id,
                        'nombre': p.user.first_name + ' ' + p.user.last_name if p.user.first_name else p.user.username,
                        'email': p.user.email,
                        'nivel': p.nivel,
                        'carrera_id': p.carrera.id if p.carrera else '',
                        'carrera_nombre': p.carrera.nombre if p.carrera else 'Sin asignar',
                        'materias_ids_json': json.dumps(materias_ids),
                        'materias_count': len(materias_ids)
                    })
                context['lista_profesores'] = lista_profesores
                
                # Cargar todas las materias para el JS de filtrado local
                all_materias = list(Materia.objects.all().values('id', 'nombre', 'carrera_id', 'año'))
                context['all_materias_data'] = all_materias
            return render(request, 'profesor.html', context)
        except Profesor.DoesNotExist:
            return render(request, 'profesor.html', {'materias': [], 'metricas': {'total_alumnos': 0, 'promedio_grupal': 0, 'tareas_pendientes': 0}})
    
    # Logica para el estudiante
    try:
        from datetime import datetime
        
        estudiante = request.user.estudiante
        
        # Obtener fecha y hora actual del servidor
        ahora = datetime.now()
        dia_actual = ahora.strftime('%A')  # Nombre del día en inglés
        hora_actual = ahora.time()
        
        # Mapeo de días en inglés a español
        dias_map = {
            'Monday': 'Lunes',
            'Tuesday': 'Martes',
            'Wednesday': 'Miércoles',
            'Thursday': 'Jueves',
            'Friday': 'Viernes',
            'Saturday': 'Sábado',
            'Sunday': 'Domingo'
        }
        dia_espanol = dias_map.get(dia_actual, dia_actual)
        
        # Obtener todos los horarios del estudiante
        horario_completo = Horario.objects.filter(
            materia__carrera=estudiante.carrera,
            materia__año=estudiante.año
        ).select_related('materia').order_by('dia', 'hora_inicio')
        
        # Calcular si hay clase ahora y cuál es la proxima
        proxima_clase = None
        tiene_clase_ahora = False
        
        # Buscar clase actual
        clases_hoy = horario_completo.filter(dia=dia_espanol)
        for clase in clases_hoy:
            if clase.hora_inicio <= hora_actual <= clase.hora_fin:
                proxima_clase = clase
                tiene_clase_ahora = True
                break
        
        # Si no hay clase ahora, buscar la próxima clase de hoy
        if not tiene_clase_ahora:
            for clase in clases_hoy:
                if clase.hora_inicio > hora_actual:
                    proxima_clase = clase
                    break
        
        # Si no hay mas clases hoy, buscar la primera clase del proximo dia
        if not proxima_clase:
            # Orden de dias de la semana
            orden_dias = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
            idx_hoy = orden_dias.index(dia_espanol) if dia_espanol in orden_dias else 0
            
            for i in range(1, 8):  # Buscar en los proximos 7 dias
                siguiente_dia = orden_dias[(idx_hoy + i) % 7]
                proxima_clase = horario_completo.filter(dia=siguiente_dia).first()
                if proxima_clase:
                    break
        
        # Notas con tendencia
        notas_raw = Nota.objects.filter(estudiante=estudiante)[:3]
        notas = []
        for n in notas_raw:
            notas.append({
                'materia': n.materia,
                'valor': n.valor,
                'tendencia': n.tendencia,
                'porcentaje': n.valor * 20 # 5 * 20 = 100
            })
        
        # Noticias categorizadas
        noticias = Noticia.objects.filter(
            Q(visible_para='todos') | 
            (
                Q(visible_para='estudiantes') & 
                (Q(carrera=estudiante.carrera) | Q(carrera__isnull=True)) & 
                (Q(anio=estudiante.año) | Q(anio__isnull=True))
            )
        ).order_by('-fecha_publicacion')[:3]

        context = {
            'proxima_clase': proxima_clase,
            'tiene_clase_ahora': tiene_clase_ahora,
            'horario_completo': horario_completo,
            'notas': notas,
            'noticias': noticias
        }
        return render(request, 'estudiante.html', context)
    except Estudiante.DoesNotExist:
        return render(request, 'estudiante.html', {})

ROLE_MAP = {
    'profesor.com': 'profesor',
    'estudiante.com': 'estudiante',
    'jefedecarrera.com': 'jefedecarrera',
}

def determinar_rol_email(email):
    dominio = (email or '').lower().split('@')[-1]
    return ROLE_MAP.get(dominio, 'estudiante')

User = get_user_model()

@login_required
def editar_noticia(request, noticia_id):
    noticia = get_object_or_404(Noticia, id=noticia_id)
    if noticia.autor.user != request.user:
        messages.error(request, 'No tienes permiso para editar esta noticia.')
        return redirect('base')
    
    if request.method == 'POST':
        form = NoticiaForm(request.POST, instance=noticia)
        if form.is_valid():
            noticia = form.save(commit=False)
            # Asegurar que el autor no cambie
            noticia.autor = request.user.profesor
            noticia.save()
            messages.success(request, 'Noticia actualizada correctamente.')
        else:
            messages.error(request, 'Error al actualizar la noticia. Verifique los campos.')
    
    return redirect('base')

@login_required
def subir_notas(request):
    if not request.user.is_profesor:
        messages.error(request, 'Acceso denegado.')
        return redirect('base')
    
    if request.method == 'POST':
        form = SubirNotasForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            carrera = form.cleaned_data['carrera']
            anio = form.cleaned_data['anio']
            materia = form.cleaned_data['materia']
            
            # Validar que materia coincida con carrera y año (Server side extra check)
            if materia.carrera != carrera or str(materia.año) != str(anio):
                messages.error(request, 'La materia no coincide con la carrera y año seleccionados.')
                return redirect('base')

            try:
                wb = openpyxl.load_workbook(archivo)
                sheet = wb.active
                
                notas_creadas = 0
                notas_actualizadas = 0
                errores = []
                
                def normalizar_texto(texto):
                    import unicodedata
                    texto_sin_tildes = ''.join(c for c in unicodedata.normalize('NFD', str(texto)) if unicodedata.category(c) != 'Mn')
                    return texto_sin_tildes.lower().strip()
                
                from django.db import transaction
                with transaction.atomic():
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        if not row or all(cell is None for cell in row): continue
                        
                        nombre_raw = row[0]
                        nota_val = row[1]
                        
                        if not nombre_raw or nota_val is None:
                            errores.append(f'Fila {row_idx}: Datos incompletos')
                            continue
                            
                        try:
                            nota_num = float(nota_val)
                            if not (0 <= nota_num <= 5):
                                errores.append(f'Fila {row_idx}: Nota {nota_num} fuera de rango (0-5)')
                                continue
                        except ValueError:
                            errores.append(f'Fila {row_idx}: Valor de nota inválido')
                            continue

                        # Buscar estudiante
                        estudiantes = Estudiante.objects.filter(carrera=carrera, año=anio).select_related('user')
                        target_name = normalizar_texto(nombre_raw)
                        estudiante = None
                        
                        for est in estudiantes:
                            # Comparar con username
                            # Usaremos username como primary, o first_name + last_name
                            n_user = normalizar_texto(est.user.username)
                            n_full = normalizar_texto(f"{est.user.first_name} {est.user.last_name}")
                            if target_name == n_user or target_name == n_full:
                                estudiante = est
                                break
                        
                        if not estudiante:
                            errores.append(f'Fila {row_idx}: Estudiante "{nombre_raw}" no encontrado.')
                            continue

                        # Procesar nota
                        defaults = {'valor': nota_num, 'tendencia': 'mantuvo'}
                        
                        # Calc tendencia
                        last_nota = Nota.objects.filter(estudiante=estudiante, materia=materia).first()
                        if last_nota:
                            if nota_num > last_nota.valor: defaults['tendencia'] = 'subio'
                            elif nota_num < last_nota.valor: defaults['tendencia'] = 'bajo'

                        obj, created = Nota.objects.update_or_create(
                            estudiante=estudiante, materia=materia,
                            defaults=defaults
                        )
                        if created: notas_creadas += 1
                        else: notas_actualizadas += 1
                
                if errores:
                    # Mostrar primeros 3 errores en toast
                    msg_err = "; ".join(errores[:3])
                    messages.warning(request, f'Procesado con errores: {msg_err} ... (Total: {len(errores)})')
                else:
                    messages.success(request, f'Éxito: {notas_creadas} creadas, {notas_actualizadas} actualizadas.')
                    
            except Exception as e:
                messages.error(request, f'Error procesando archivo: {str(e)}')
        else:
            messages.error(request, 'Formulario inválido. Verifique los campos.')

    return redirect('base')


@login_required
def get_materias_api(request):
    """API para obtener materias filtradas por carrera y año"""
    try:
        carrera_id = request.GET.get('carrera_id')
        anio = request.GET.get('anio')
        
        if not carrera_id or not anio:
            return JsonResponse({'materias': []})
        
        materias = Materia.objects.filter(
            carrera_id=carrera_id,
            año=anio
        ).values('id', 'nombre')
        
        return JsonResponse({'materias': list(materias)})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def asignar_profesor_api(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    # Verificar permisos
    if not hasattr(request.user, 'profesor') or request.user.profesor.nivel != 'jefe de carrera':
        return JsonResponse({'error': 'No tienes permisos para realizar esta acción'}, status=403)
        
    try:
        data = json.loads(request.body)
        profesor_id = data.get('profesor_id')
        carrera_id = data.get('carrera_id')
        materias_ids = data.get('materias_ids', [])
        
        if not profesor_id:
            return JsonResponse({'error': 'Falta ID de profesor'})
            
        profesor = Profesor.objects.get(id=profesor_id)
        
        # Asignar Carrera
        if carrera_id:
            profesor.carrera_id = carrera_id
        else:
            profesor.carrera = None
        profesor.save()
        
        # Asignar Materias
        # Aseguramos que los IDs sean enteros y existan
        if materias_ids:
             materias_ids = [int(m_id) for m_id in materias_ids]
             profesor.materias.set(materias_ids)
        else:
             profesor.materias.clear()
        
        return JsonResponse({'success': True})
        
    except Profesor.DoesNotExist:
        return JsonResponse({'error': 'Profesor no encontrado'}, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def subir_horarios(request):
    if not request.user.is_profesor:
        messages.error(request, 'Acceso denegado')
        return redirect('base')
    
    if request.user.profesor.nivel != 'jefe de carrera':
        messages.error(request, 'Solo los jefes de carrera pueden subir horarios')
        return redirect('base')
    
    if request.method == 'POST':
        form = SubirHorariosForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            carrera = form.cleaned_data['carrera']
            anio = form.cleaned_data['anio']
            reemplazar = form.cleaned_data['reemplazar']
            
            try:
                from django.conf import settings
                from django.db import transaction
                from datetime import datetime, time
                import openpyxl

                wb = openpyxl.load_workbook(archivo)
                sheet = wb.active
                
                horarios_creados = 0
                horarios_actualizados = 0
                errores = []

                if reemplazar:
                   # Eliminar horarios de esa carrera y año
                   Horario.objects.filter(materia__carrera=carrera, materia__año=anio).delete()

                def parsear_hora(valor):
                    if isinstance(valor, time): return valor
                    if isinstance(valor, str):
                        try:
                            parts = valor.strip().split(':')
                            return time(int(parts[0]), int(parts[1]))
                        except: pass
                    return None
                    
                def normalizar_texto(texto):
                    import unicodedata
                    return ''.join(c for c in unicodedata.normalize('NFD', str(texto)) if unicodedata.category(c) != 'Mn').lower().strip()

                with transaction.atomic():
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        if not row or all(cell is None for cell in row): continue
                        
                        materia_nombre = row[0]
                        dia = row[1]
                        hora_ini_raw = row[2]
                        hora_fin_raw = row[3]
                        aula = row[4] if len(row) > 4 else ''
                        
                        if not all([materia_nombre, dia, hora_ini_raw, hora_fin_raw]):
                             errores.append(f"Fila {row_idx}: Datos incompletos")
                             continue
                             
                        hora_inicio = parsear_hora(hora_ini_raw)
                        hora_fin = parsear_hora(hora_fin_raw)
                        
                        if not hora_inicio or not hora_fin:
                            errores.append(f"Fila {row_idx}: Hora inválida")
                            continue
                            
                        # Buscar materia
                        materia = Materia.objects.filter(carrera=carrera, año=anio, nombre__iexact=materia_nombre.strip()).first()
                        if not materia:
                             errores.append(f"Fila {row_idx}: Materia '{materia_nombre}' no encontrada en carrera/año")
                             continue
                        
                        Horario.objects.update_or_create(
                            materia=materia,
                            dia=dia.capitalize(),
                            hora_inicio=hora_inicio,
                            defaults={'hora_fin': hora_fin, 'aula': aula}
                        )
                        horarios_creados += 1

                if errores:
                     msg = "; ".join(errores[:3])
                     messages.warning(request, f"Procesado con errores: {msg} ...")
                else:
                     messages.success(request, f"Horarios procesados correctamente ({horarios_creados} registros).")
                     
            except Exception as e:
                messages.error(request, f"Error procesando archivo: {str(e)}")
        else:
            messages.error(request, "Formulario inválido")
            
    return redirect('base')

@login_required
def obtener_estudiantes_notas_api(request, materia_id):
    """API para obtener estudiantes y sus notas de una materia específica"""
    try:
        # Verificar que el usuario sea profesor
        if not request.user.is_profesor:
            return JsonResponse({'error': 'Acceso denegado'}, status=403)
        
        # Obtener la materia
        try:
            materia = Materia.objects.get(id=materia_id)
        except Materia.DoesNotExist:
            return JsonResponse({'error': 'Materia no encontrada'}, status=404)
        
        # Obtener todos los estudiantes de la carrera y año de la materia
        estudiantes = Estudiante.objects.filter(
            carrera=materia.carrera,
            año=materia.año
        ).select_related('user').order_by('user__username')
        
        # Crear lista de estudiantes con sus notas
        estudiantes_data = []
        for estudiante in estudiantes:
            # Buscar si tiene nota en esta materia
            try:
                nota = Nota.objects.get(estudiante=estudiante, materia=materia)
                nota_valor = nota.valor
                tendencia = nota.tendencia
            except Nota.DoesNotExist:
                nota_valor = None
                tendencia = None
            
            estudiantes_data.append({
                'id': estudiante.id,
                'nombre': estudiante.user.username,
                'nota': nota_valor,
                'tendencia': tendencia
            })
        
        return JsonResponse({
            'estudiantes': estudiantes_data,
            'materia': materia.nombre
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'trace': traceback.format_exc()}, status=500)

@login_required
def crear_noticia(request):
    if request.method == 'POST' and request.user.is_profesor:
        form = NoticiaForm(request.POST)
        if form.is_valid():
            noticia = form.save(commit=False)
            noticia.autor = request.user.profesor
            noticia.save()
            messages.success(request, 'Noticia publicada exitosamente.')
        else:
            messages.error(request, 'Error al publicar noticia. Verifique los campos.')
    return redirect('base')

@login_required
def borrar_noticia(request, noticia_id):
    if request.method == 'POST' and request.user.is_profesor:
        noticia = Noticia.objects.filter(id=noticia_id, autor=request.user.profesor).first()
        if noticia:
            noticia.delete()
            messages.success(request, 'Noticia eliminada.')
        else:
            messages.error(request, 'No tienes permiso para eliminar esta noticia.')
    return redirect('base')
